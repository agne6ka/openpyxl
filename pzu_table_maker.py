from openpyxl import load_workbook
import os


def xlsx_to_html():
    wb = load_workbook('/home/aga/Desktop/Zeszyt2.xlsx')
    for sheet in wb:
        content = build_content(sheet)
        name = os.path.relpath(os.path.join('results', sheet.title + '.html'))
        write_to_file(content, name)


def build_content(sheet):
    prefix = '<table class="zebra"><thead>' \
             '<tr><th>Nazwa</th></tr>' \
             '<tr><th>Adres</th></tr>' \
             '</thead><tbody>'
    suffix = '</tbody></table>'
    # body_rows = [u'<tr class="odd">{}</tr>'.format(''.join([u'<td>{}</td>'.format(cell.value or '') for cell in row if cell.row]))
    #              for row in ws.iter_rows()]
    body_rows = [u'<tr class="{cls}">{body}</tr>'.format(
        cls='even' if row[0].row % 2 == 0 else 'odd',
        body=''.join((u'<td>{}</td>'.format(cell.value or '') for cell in row if cell.row and cell.column < 'C'))
    ) for row in sheet.iter_rows()]
    return u'{}{}{}'.format(prefix, '\n'.join(body_rows), suffix)


def write_to_file(content, filename):
    with open(filename, 'w') as f:
        f.write(content.encode('utf-8'))

xlsx_to_html()
