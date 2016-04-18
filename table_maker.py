from openpyxl import load_workbook


def xlsx_to_html():
    wb = load_workbook('/home/aga/Desktop/Good_things.xlsx')
    ws = wb.active

    prefix = '<table class="zebra"><thead>' \
             '<tr><th>Nazwa</th></tr>' \
             '<tr><th>Adres</th></tr>' \
             '</thead><tbody>'
    suffix = '</tbody></table>'
    # body_rows = [u'<tr class="odd">{}</tr>'.format(''.join([u'<td>{}</td>'.format(cell.value or '') for cell in row if cell.row]))
    #              for row in ws.iter_rows()]
    body_rows = [u'<tr class="{cls}">{body}</tr>'.format(
        cls='even' if row[0].row % 2 == 0 else 'odd',
        body=''.join([u'<td>{}</td>'.format(cell.value or '') for cell in row if cell.row])
    ) for row in ws.iter_rows()]
    return u'{}{}{}'.format(prefix, '\n'.join(body_rows), suffix)


def write_to_file(content, filename):
    with open(filename, 'w') as f:
        f.write(content.encode('utf-8'))


html = xlsx_to_html()
write_to_file(html, 'test.html')